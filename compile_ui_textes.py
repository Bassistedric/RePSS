import openpyxl, json

wb = openpyxl.load_workbook("/mnt/user-data/outputs/RePSS_UI_Textes_maitre.xlsx", data_only=True)
ws = wb["UI_Textes"]

headers = [c.value for c in ws[1]]
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    rows.append(dict(zip(headers, row)))

for lang in ("fr", "en", "nl"):
    ui = {}
    missing = []
    for r in rows:
        val = r.get(lang)
        if not val or not str(val).strip():
            missing.append(r["id"])
            val = r.get("fr", "")  # repli sur le français si une traduction manque
        ui[r["id"]] = val

    data = {"ui": ui, "risques": []}  # catalogue de risques traduit à ajouter plus tard

    with open(f"/mnt/user-data/outputs/{lang}.json", "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"{lang}.json — {len(ui)} clés" + (f" — {len(missing)} repli sur le français: {missing}" if missing else " — complet"))
