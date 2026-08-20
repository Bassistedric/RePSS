import openpyxl, json

# Compile le catalogue d'analyse de risque en 3 langues (fr/en/nl), à partir de
# 3 classeurs maîtres parallèles (même structure/ids, texte traduit) :
#   RePSS_Analyse_Risques.xlsx (fr, source de vérité pour la structure et les
#   valeurs Kinney numériques) / RePSS_Analyse_Risques_EN.xlsx / _NL.xlsx.
#
# `niveauCode` (acceptable/attention/correction/immediate/arret) est calculé une
# seule fois à partir du texte `eval_ini_niveau`/`eval_res_niveau` du classeur
# FRANÇAIS (seule langue de référence pour ce calcul) puis reporté à l'identique
# sur les 3 langues via l'id de la ligne — ça évite à `couleurNiveau()` (lib/
# kinney.js) de dépendre d'une correspondance de texte traduit, fragile.
NIVEAU_CODE_PAR_TEXTE_FR = {
    "Acceptable": "acceptable",
    "Attention requise": "attention",
    "Correction nécessaire": "correction",
    "Mesure immédiate": "immediate",
    "Envisager l'arrêt": "arret",
}

WORKBOOKS = {
    "fr": "RePSS_Analyse_Risques.xlsx",
    "en": "RePSS_Analyse_Risques_EN.xlsx",
    "nl": "RePSS_Analyse_Risques_NL.xlsx",
}

RISQUE_HEADERS = [
    "id", "parent", "sourceDanger", "risques",
    "eval_ini_probabilite_texte", "eval_ini_probabilite_valeur",
    "eval_ini_exposition_texte", "eval_ini_exposition_valeur",
    "eval_ini_gravite_texte", "eval_ini_gravite_valeur",
    "eval_ini_score", "eval_ini_niveau",
    "mesures_prevention",
    "eval_res_probabilite_texte", "eval_res_probabilite_valeur",
    "eval_res_exposition_texte", "eval_res_exposition_valeur",
    "eval_res_gravite_texte", "eval_res_gravite_valeur",
    "eval_res_score", "eval_res_niveau",
]


def rows_of(wb, sheet, headers):
    ws = wb[sheet]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        out.append(dict(zip(headers, row)))
    return out


def niveau_codes_from_french():
    wb = openpyxl.load_workbook(WORKBOOKS["fr"], data_only=True)
    risques_raw = rows_of(wb, "LignesRisque", RISQUE_HEADERS)
    codes = {}
    for r in risques_raw:
        codes[r["id"]] = {
            "ini": NIVEAU_CODE_PAR_TEXTE_FR.get(r["eval_ini_niveau"]),
            "res": NIVEAU_CODE_PAR_TEXTE_FR.get(r["eval_res_niveau"]),
        }
    return codes


def compile_lang(lang, niveau_codes):
    wb = openpyxl.load_workbook(WORKBOOKS[lang], data_only=True)

    categories = rows_of(wb, "Categories", ["id", "code_origine", "fr", "corps_metier", "groupe"])
    souscategories = rows_of(wb, "SousCategories", ["id", "code_origine", "fr", "parent"])
    activites = rows_of(wb, "Activites", ["id", "code_origine", "fr", "parent", "granularite", "confiance", "notes"])
    risques_raw = rows_of(wb, "LignesRisque", RISQUE_HEADERS)

    risques = []
    for r in risques_raw:
        codes = niveau_codes.get(r["id"], {})
        risques.append({
            "id": r["id"], "parent": r["parent"],
            "sourceDanger": r["sourceDanger"], "risques": r["risques"],
            "evaluationInitiale": {
                "probabilite": {"texte": r["eval_ini_probabilite_texte"], "valeur": r["eval_ini_probabilite_valeur"]},
                "exposition": {"texte": r["eval_ini_exposition_texte"], "valeur": r["eval_ini_exposition_valeur"]},
                "gravite": {"texte": r["eval_ini_gravite_texte"], "valeur": r["eval_ini_gravite_valeur"]},
                "score": r["eval_ini_score"], "niveau": r["eval_ini_niveau"], "niveauCode": codes.get("ini"),
            },
            "mesuresPrevention": r["mesures_prevention"],
            "evaluationResiduelle": {
                "probabilite": {"texte": r["eval_res_probabilite_texte"], "valeur": r["eval_res_probabilite_valeur"]},
                "exposition": {"texte": r["eval_res_exposition_texte"], "valeur": r["eval_res_exposition_valeur"]},
                "gravite": {"texte": r["eval_res_gravite_texte"], "valeur": r["eval_res_gravite_valeur"]},
                "score": r["eval_res_score"], "niveau": r["eval_res_niveau"], "niveauCode": codes.get("res"),
            },
        })

    data = {
        "categories": categories,
        "sousCategories": souscategories,
        "activites": activites,
        "lignesRisque": risques,
    }

    suffix = "" if lang == "fr" else f"_{lang}"
    out_path = f"app/public/content-pack/catalogue_risques{suffix}.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"{out_path} — {len(categories)} catégories, {len(souscategories)} sous-cat, "
          f"{len(activites)} activités, {len(risques)} lignes de risque")


if __name__ == "__main__":
    codes = niveau_codes_from_french()
    for lang in ("fr", "en", "nl"):
        compile_lang(lang, codes)
